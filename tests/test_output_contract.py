import csv
import tempfile
import unittest
from pathlib import Path

import openpyxl

from working_paper_review_engine.output_contract import (
    COMMENT_MATRIX_COLUMNS,
    COMMENT_MATRIX_SHEET_NAME,
    COMMENT_TYPE_VALUES,
    CommentMatrixRow,
    infer_comment_type,
    rows_from_comments,
    rows_to_csv,
    rows_to_xlsx,
)


class OutputContractTests(unittest.TestCase):
    def test_column_order_matches_canonical_contract(self) -> None:
        """COMMENT_MATRIX_COLUMNS must exactly match the spectrum-systems 13-column contract."""
        expected = [
            "Comment Number",
            "Reviewer Initials",
            "Agency",
            "Report Version",
            "Section",
            "Page",
            "Line",
            "Comment Type: Editorial/Grammar, Clarification, Technical",
            "Agency Notes",
            "Agency Suggested Text Change",
            "NTIA Comments",
            "Comment Disposition",
            "Resolution",
        ]
        self.assertEqual(COMMENT_MATRIX_COLUMNS, expected)

    def test_column_count_is_exactly_thirteen(self) -> None:
        """No extra visible columns beyond the 13 canonical ones."""
        self.assertEqual(len(COMMENT_MATRIX_COLUMNS), 13)

    def test_comment_type_values_are_canonical(self) -> None:
        """Allowed Comment Type values must match the spectrum-systems contract exactly."""
        self.assertEqual(
            sorted(COMMENT_TYPE_VALUES),
            sorted(["Editorial/Grammar", "Clarification", "Technical"]),
        )

    def test_sheet_name_is_canonical(self) -> None:
        """Sheet name must be 'Comment Resolution Matrix' per the spectrum-systems contract."""
        self.assertEqual(COMMENT_MATRIX_SHEET_NAME, "Comment Resolution Matrix")

    def test_rows_serialize_to_csv_and_xlsx(self) -> None:
        comment = {
            "comment_number": 7,
            "reviewer_persona_id": "FCC policy analyst",
            "comment_text": "Align definitions of primary/secondary allocation with FCC Part 2 wording.",
            "suggested_revision": "Use FCC Part 2 definitions verbatim in Section 1.1.",
            "rationale": "Anchored to p.3 lines 5-8; current phrasing deviates from CFR text.",
            "anchor_text": "Secondary users shall not cause harmful interference...",
            "section_id": "1.1",
            "page_reference": "3",
            "line_reference": "5-8",
            "comment_type": "Clarification",
            "comment_disposition": "For consideration",
            "resolution": "Update wording and cite CFR Part 2.104(d).",
            "confidence": 0.64,
            "heat_level": "minor",
        }
        rows = rows_from_comments([comment], report_version="REV-A")
        tmpdir = Path(tempfile.mkdtemp())

        csv_path = rows_to_csv(rows, tmpdir / "matrix.csv")
        xlsx_path = rows_to_xlsx(rows, tmpdir / "matrix.xlsx")

        with csv_path.open() as handle:
            reader = csv.reader(handle)
            header = next(reader)
            row = next(reader)
        self.assertEqual(header, COMMENT_MATRIX_COLUMNS)
        agency_notes_idx = COMMENT_MATRIX_COLUMNS.index("Agency Notes")
        suggested_idx = COMMENT_MATRIX_COLUMNS.index("Agency Suggested Text Change")
        comment_type_idx = COMMENT_MATRIX_COLUMNS.index(
            "Comment Type: Editorial/Grammar, Clarification, Technical"
        )
        self.assertEqual(row[agency_notes_idx], comment["comment_text"])
        self.assertEqual(row[suggested_idx], comment["suggested_revision"])
        self.assertEqual(row[comment_type_idx], "Clarification")

        workbook = openpyxl.load_workbook(xlsx_path)
        self.assertIn(COMMENT_MATRIX_SHEET_NAME, workbook.sheetnames)
        sheet = workbook[COMMENT_MATRIX_SHEET_NAME]
        header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(header_row, COMMENT_MATRIX_COLUMNS)
        value_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        disposition_idx = COMMENT_MATRIX_COLUMNS.index("Comment Disposition")
        self.assertEqual(value_row[disposition_idx], comment["comment_disposition"])

    def test_no_extra_columns_in_ordered_mapping(self) -> None:
        """to_ordered_mapping must emit exactly the 13 canonical columns and nothing else."""
        row = CommentMatrixRow(
            comment_number=1,
            persona_name="DoD spectrum engineer",
            agency_notes="Sample",
        ).to_ordered_mapping()
        self.assertEqual(list(row.keys()), COMMENT_MATRIX_COLUMNS)

    def test_infer_comment_type_technical(self) -> None:
        self.assertEqual(infer_comment_type("technical sufficiency"), "Technical")
        self.assertEqual(infer_comment_type("reproducibility"), "Technical")
        self.assertEqual(infer_comment_type("technical"), "Technical")

    def test_infer_comment_type_clarification(self) -> None:
        self.assertEqual(infer_comment_type("regulatory ambiguity"), "Clarification")
        self.assertEqual(infer_comment_type("clarity"), "Clarification")
        self.assertEqual(infer_comment_type("clarification"), "Clarification")

    def test_infer_comment_type_editorial(self) -> None:
        self.assertEqual(infer_comment_type("inconsistent terminology"), "Editorial/Grammar")
        self.assertEqual(infer_comment_type("editorial"), "Editorial/Grammar")

    def test_infer_comment_type_blank_on_unknown(self) -> None:
        """Unknown category with no severity should return empty string (not fabricate a type)."""
        self.assertEqual(infer_comment_type(""), "")
        self.assertEqual(infer_comment_type("unknown-random-category"), "")

    def test_infer_comment_type_fallback_by_severity(self) -> None:
        self.assertEqual(infer_comment_type("", "critical"), "Technical")
        self.assertEqual(infer_comment_type("", "major"), "Technical")
        self.assertEqual(infer_comment_type("", "minor"), "Clarification")
        self.assertEqual(infer_comment_type("", "info"), "Clarification")


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
